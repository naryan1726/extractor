import streamlit as st
import os
import re
import json
import pandas as pd
import time
import io
from datetime import datetime
from typing import Dict, List, Optional
from dataclasses import dataclass, field

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# For PDF processing
import pdfplumber

# For Excel generation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Groq API
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False
    st.error("Groq library not installed. Installing required packages...")
    os.system("pip install groq")
    from groq import Groq
    GROQ_AVAILABLE = True

@dataclass
class ContactInfo:
    """Contact information data structure"""
    name: str = ""
    phone: str = ""
    email: str = ""

@dataclass
class F913ExtractionResult:
    """F913 extraction result with metadata"""
    file_name: str
    file_path: str = ""  # Not used in Streamlit version
    tenant: ContactInfo = field(default_factory=ContactInfo)
    landlord: ContactInfo = field(default_factory=ContactInfo)
    extraction_method: str = "Unknown"
    model_used: str = ""
    confidence_score: float = 0.0
    processing_time: float = 0.0
    processed_datetime: str = ""

class GroqF913Extractor:
    """High-performance F913 extractor using latest Groq models"""
    
    # Latest Groq models (Jan 2025)
    MODELS = {
        "premium": "llama-3.3-70b-versatile",      # Best accuracy, 128K context
        "fast": "llama-3.3-70b-specdec",           # Fastest 70B, speculative decoding
        "instant": "llama-3.1-8b-instant",         # Quick responses, lower cost
        "fallback": "llama-3.1-70b-versatile"      # Backup option
    }
    
    def __init__(self, api_key: Optional[str] = None):
        """Initialize with Groq API key"""
        self.api_key = api_key
        self.client = None
        self.default_model = self.MODELS["premium"]
        
        if not self.api_key:
            st.error("❌ GROQ_API_KEY not found! Please enter your API key in the sidebar.")
            return
        
        try:
            self.client = Groq(api_key=self.api_key)
            # Test connection (minimally)
            self.client.chat.completions.create(
                model=self.MODELS["instant"],
                messages=[{"role": "user", "content": "Hi"}],
                max_tokens=5
            )
            st.success("✅ Groq API connected successfully")
        except Exception as e:
            st.error(f"❌ Groq initialization failed: {str(e)}")
            self.client = None
    
    def extract_f913_data(self, pdf_content: bytes, filename: str) -> Optional[F913ExtractionResult]:
        """Extract data from F913 PDF using Groq"""
        start_time = time.time()
        
        # Create result with timestamp
        processed_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        result = F913ExtractionResult(
            file_name=filename,
            tenant=ContactInfo(),
            landlord=ContactInfo(),
            processed_datetime=processed_datetime
        )
        
        try:
            # Extract PDF text
            signature_text, full_text = self._extract_pdf_text(pdf_content)
            
            if not signature_text and not full_text:
                st.error("❌ No text extracted from PDF")
                return None
            
            # Use signature page if available, otherwise full text
            text_to_analyze = signature_text or full_text
            
            # Truncate text for API efficiency
            if len(text_to_analyze) > 8000:
                text_to_analyze = text_to_analyze[:8000] + "..."
            
            # Try extraction with different models
            for model_type, model_id in self.MODELS.items():
                try:
                    st.write(f"🤖 Trying {model_type} model: {model_id}")
                    
                    extracted_data = self._extract_with_groq_model(
                        text_to_analyze, model_id, model_type
                    )
                    
                    if extracted_data:
                        # Update result
                        result.tenant = ContactInfo(**extracted_data.get("tenant", {}))
                        result.landlord = ContactInfo(**extracted_data.get("landlord", {}))
                        result.extraction_method = "Groq API"
                        result.model_used = f"{model_type} ({model_id})"
                        result.confidence_score = extracted_data.get("confidence", 0.8)
                        result.processing_time = time.time() - start_time
                        
                        # Validate and correct potential mix-ups
                        self._post_process_validation(result, text_to_analyze)
                        
                        # Check if we got useful data
                        if result.tenant.name or result.landlord.name:
                            return result
                
                except Exception as e:
                    st.warning(f"⚠️ {model_type} model failed: {str(e)}")
                    continue
            
            # If all models failed, try regex fallback
            st.write("🔍 Falling back to regex extraction")
            regex_result = self._extract_with_regex(text_to_analyze, result)
            if regex_result:
                regex_result.processing_time = time.time() - start_time
                return regex_result
            
            return None
            
        except Exception as e:
            st.error(f"❌ Error processing {filename}: {str(e)}")
            return None
    
    def _extract_pdf_text(self, pdf_content: bytes) -> tuple:
        """Extract text from PDF, prioritizing signature page"""
        signature_text = ""
        full_text = ""
        
        try:
            with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
                pages = pdf.pages
                
                # Check last page for signature
                if pages:
                    last_page_text = pages[-1].extract_text() or ""
                    if any(marker in last_page_text for marker in 
                           ["IN WITNESS WHEREOF", "Tenant's Signature", "Landlord's Signature"]):
                        signature_text = last_page_text
                        st.write("📋 Found signature page")
                
                # Get full text as fallback
                for page in pages:
                    page_text = page.extract_text() or ""
                    full_text += page_text + "\n"
                    
                    # Look for signature page in all pages if not found
                    if not signature_text and "IN WITNESS WHEREOF" in page_text:
                        signature_text = page_text
            
            return signature_text.strip(), full_text.strip()
            
        except Exception as e:
            st.error(f"⚠️ PDF extraction error: {str(e)}")
            return "", ""
    
    def _extract_with_groq_model(self, text: str, model_id: str, model_type: str) -> Optional[dict]:
        """Extract data using specific Groq model"""
        
        # Adjust prompt based on model capability
        if model_type in ["premium", "fast"]:
            # Detailed prompt for powerful models with improved context separation
            prompt = f"""You are an expert at extracting contact information from F913 Lease for Residential Property documents.

TASK: Analyze this F913 document text and extract tenant and landlord contact information.

IMPORTANT INSTRUCTIONS:
1. Look for the signature page section with "IN WITNESS WHEREOF"
2. Find TWO distinct sections:
   - "Tenant's Signature" section (left side) - contains TENANT information
   - "Landlord's Signature" section (right side) - contains LANDLORD information
3. Each section has "Print or Type Name", phone number, and email address fields
4. DO NOT mix up the two parties - tenant info goes to tenant, landlord info goes to landlord
5. If you see names like "Sam Kumar" and "Apuroop", determine which is tenant vs landlord based on their position in the document

CRITICAL: Pay attention to the document structure - there are separate fields for each party.

Return ONLY a valid JSON object in this exact format:
{{
    "tenant": {{
        "name": "Full Name Here",
        "phone": "1234567890",
        "email": "email@domain.com"
    }},
    "landlord": {{
        "name": "Full Name Here", 
        "phone": "1234567890",
        "email": "email@domain.com"
    }},
    "confidence": 0.95
}}

Rules:
- Use empty strings "" for any missing information
- Phone numbers should be 10 digits only
- DO NOT switch tenant and landlord information
- First name/phone/email usually belongs to tenant, second to landlord
- Confidence should be 0.0 to 1.0
- Return ONLY the JSON, no explanation

Document text:
{text[:4000]}"""
        
        else:
            # Simplified prompt for smaller models
            prompt = f"""Extract tenant and landlord info from F913 lease document.

IMPORTANT: Don't mix up tenant and landlord info. First person is usually tenant, second is landlord.

Return JSON only:
{{
    "tenant": {{"name": "", "phone": "", "email": ""}},
    "landlord": {{"name": "", "phone": "", "email": ""}},
    "confidence": 0.8
}}

Document: {text[:2500]}"""
        
        try:
            # Configure based on model type
            if model_type == "instant":
                max_tokens = 300
                temperature = 0.1
            elif model_type == "fast":
                max_tokens = 400  
                temperature = 0.05
            else:  # premium
                max_tokens = 500
                temperature = 0.02
            
            response = self.client.chat.completions.create(
                model=model_id,
                messages=[
                    {
                        "role": "system", 
                        "content": "You are a precise document data extraction specialist. Return only valid JSON responses."
                    },
                    {
                        "role": "user", 
                        "content": prompt
                    }
                ],
                max_tokens=max_tokens,
                temperature=temperature,
                top_p=0.9
            )
            
            content = response.choices[0].message.content.strip()
            
            # Clean JSON response
            content = self._clean_json_response(content)
            
            # Parse JSON
            data = json.loads(content)
            
            # Validate structure
            if self._validate_extraction_data(data):
                return data
            else:
                st.warning("⚠️ Invalid data structure from model")
                return None
                
        except json.JSONDecodeError as e:
            st.warning(f"⚠️ JSON parse error: {str(e)}")
            # Try to extract JSON from response
            return self._extract_json_from_text(content)
        except Exception as e:
            st.warning(f"⚠️ API error: {str(e)}")
            return None
    
    def _clean_json_response(self, content: str) -> str:
        """Clean and extract JSON from model response"""
        # Remove markdown formatting
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        
        # Find JSON object
        start = content.find('{')
        end = content.rfind('}') + 1
        
        if start >= 0 and end > start:
            return content[start:end]
        
        return content.strip()
    
    def _extract_json_from_text(self, text: str) -> Optional[dict]:
        """Extract JSON from text using regex"""
        try:
            # Find JSON-like structure
            json_pattern = r'\{[^{}]*\{[^{}]*\}[^{}]*\{[^{}]*\}[^{}]*\}'
            match = re.search(json_pattern, text, re.DOTALL)
            
            if match:
                return json.loads(match.group())
            
            # Fallback: try to parse any {.*} structure
            simple_pattern = r'\{.*\}'
            match = re.search(simple_pattern, text, re.DOTALL)
            
            if match:
                return json.loads(match.group())
            
            return None
        except:
            return None
    
    def _validate_extraction_data(self, data: dict) -> bool:
        """Validate extracted data structure"""
        try:
            return (
                isinstance(data, dict) and
                "tenant" in data and
                "landlord" in data and
                isinstance(data["tenant"], dict) and
                isinstance(data["landlord"], dict)
            )
        except:
            return False
    
    def _extract_with_regex(self, text: str, result: F913ExtractionResult) -> Optional[F913ExtractionResult]:
        """Fallback regex extraction with improved party separation"""
        try:
            # Clean and normalize text
            normalized_text = re.sub(r'\s+', ' ', text)
            
            # Try to split the signature section into tenant and landlord parts
            tenant_section = ""
            landlord_section = ""
            
            # Look for section separators
            if "Tenant's Signature" in text and "Landlord's Signature" in text:
                # Split by landlord signature to separate sections
                parts = text.split("Landlord's Signature")
                if len(parts) >= 2:
                    tenant_section = parts[0]  # Everything before landlord section
                    landlord_section = "Landlord's Signature" + parts[1]  # Landlord section
                    st.write("📋 Found separate tenant and landlord sections")
            
            # F913-specific name patterns with section awareness
            if tenant_section and landlord_section:
                # Extract from separate sections
                tenant_name_patterns = [
                    r'Tenant[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                    r'1\s+Tenant[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print',
                ]
                
                landlord_name_patterns = [
                    r'Landlord[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                    r'1\s+Landlord[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print',
                ]
                
                # Extract tenant name from tenant section
                for pattern in tenant_name_patterns:
                    match = re.search(pattern, tenant_section, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.tenant.name = name
                            st.write(f"👤 Tenant name from section: {name}")
                            break
                
                # Extract landlord name from landlord section
                for pattern in landlord_name_patterns:
                    match = re.search(pattern, landlord_section, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.landlord.name = name
                            st.write(f"🏠 Landlord name from section: {name}")
                            break
                
                # Extract phones with context
                tenant_phones = re.findall(r'(\d{10})', tenant_section)
                landlord_phones = re.findall(r'(\d{10})', landlord_section)
                
                if tenant_phones:
                    result.tenant.phone = tenant_phones[0]
                    st.write(f"📞 Tenant phone from section: {tenant_phones[0]}")
                
                if landlord_phones:
                    result.landlord.phone = landlord_phones[0]
                    st.write(f"📞 Landlord phone from section: {landlord_phones[0]}")
                
                # Extract emails with context
                tenant_emails = re.findall(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', tenant_section)
                landlord_emails = re.findall(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', landlord_section)
                
                if tenant_emails:
                    result.tenant.email = tenant_emails[0]
                    st.write(f"📧 Tenant email from section: {tenant_emails[0]}")
                
                if landlord_emails:
                    result.landlord.email = landlord_emails[0]
                    st.write(f"📧 Landlord email from section: {landlord_emails[0]}")
            
            else:
                # Fallback to original patterns but with better logic
                st.write("⚠️ Could not separate sections, using positional extraction")
                
                patterns = {
                    'tenant_name': [
                        r'1\s+Tenant[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                        r'Tenant[\'s]*\s+Signature\s+([A-Za-z\s]{3,40}?)\s+Print\s+or\s+Type\s+Name',
                    ],
                    'landlord_name': [
                        r'1\s+Landlord[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                        r'Landlord[\'s]*\s+Signature\s+([A-Za-z\s]{3,40}?)\s+Print\s+or\s+Type\s+Name',
                    ]
                }
                
                # Extract names using order-aware patterns
                for pattern_list in patterns['tenant_name']:
                    match = re.search(pattern_list, normalized_text, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.tenant.name = name
                            st.write(f"👤 Tenant name: {name}")
                            break
                
                for pattern_list in patterns['landlord_name']:
                    match = re.search(pattern_list, normalized_text, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.landlord.name = name
                            st.write(f"🏠 Landlord name: {name}")
                            break
                
                # Extract phones and emails with positional logic
                all_phones = re.findall(r'(\d{10})', normalized_text)
                all_emails = re.findall(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', normalized_text)
                
                # Assign based on order (first phone/email to tenant, second to landlord)
                if all_phones:
                    if len(all_phones) >= 1:
                        result.tenant.phone = all_phones[0]
                        st.write(f"📞 Tenant phone (position 1): {all_phones[0]}")
                    if len(all_phones) >= 2:
                        result.landlord.phone = all_phones[1]
                        st.write(f"📞 Landlord phone (position 2): {all_phones[1]}")
                
                if all_emails:
                    if len(all_emails) >= 1:
                        result.tenant.email = all_emails[0]
                        st.write(f"📧 Tenant email (position 1): {all_emails[0]}")
                    if len(all_emails) >= 2:
                        result.landlord.email = all_emails[1]
                        st.write(f"📧 Landlord email (position 2): {all_emails[1]}")
            
            # Validation: Check if we mixed up names (common names heuristic)
            self._validate_and_correct_assignment(result)
            
            result.extraction_method = "Enhanced Regex"
            result.model_used = "Pattern Matching with Section Separation"
            result.confidence_score = 0.7 if (result.tenant.name or result.landlord.name) else 0.3
            
            return result if (result.tenant.name or result.landlord.name or result.tenant.phone or result.tenant.email) else None
            
        except Exception as e:
            st.error(f"⚠️ Regex extraction error: {str(e)}")
            return None
    
    def _validate_and_correct_assignment(self, result: F913ExtractionResult):
        """Validate and correct tenant/landlord assignment if needed"""
        try:
            # Simple heuristics to detect if assignments might be wrong
            # This is based on common patterns in F913 documents
            
            tenant_name = result.tenant.name.lower() if result.tenant.name else ""
            landlord_name = result.landlord.name.lower() if result.landlord.name else ""
            
            # Check if we have both names
            if tenant_name and landlord_name:
                # If landlord name appears to be a first name only while tenant has full name,
                # they might be swapped
                tenant_parts = tenant_name.split()
                landlord_parts = landlord_name.split()
                
                # Heuristic: If "landlord" has only one word and "tenant" has multiple,
                # and landlord word appears in tenant's name, they might be swapped
                if (len(landlord_parts) == 1 and len(tenant_parts) > 1 and 
                    landlord_parts[0] in tenant_parts):
                    st.warning("⚠️ Detected possible name mix-up, checking...")
                    
                    # Don't auto-correct as it might be wrong, just flag
                    result.confidence_score = max(0.1, result.confidence_score - 0.2)
                    st.write("📝 Lowered confidence due to potential mix-up")
            
        except Exception:
            # If validation fails, don't crash
            pass
    
    def _post_process_validation(self, result: F913ExtractionResult, original_text: str):
        """Post-process validation to detect and fix common mix-ups"""
        try:
            # Check for obvious mix-ups by analyzing the original text structure
            if not (result.tenant.name and result.landlord.name):
                return  # Can't validate if we don't have both names
            
            # Look for positional clues in the original text
            tenant_name_pos = original_text.find(result.tenant.name)
            landlord_name_pos = original_text.find(result.landlord.name)
            
            # Look for section indicators
            tenant_section_pos = original_text.find("Tenant's Signature")
            landlord_section_pos = original_text.find("Landlord's Signature")
            
            if tenant_section_pos != -1 and landlord_section_pos != -1:
                # Check if names are in wrong sections
                tenant_to_tenant_section = abs(tenant_name_pos - tenant_section_pos) if tenant_name_pos != -1 else float('inf')
                tenant_to_landlord_section = abs(tenant_name_pos - landlord_section_pos) if tenant_name_pos != -1 else float('inf')
                
                landlord_to_tenant_section = abs(landlord_name_pos - tenant_section_pos) if landlord_name_pos != -1 else float('inf')
                landlord_to_landlord_section = abs(landlord_name_pos - landlord_section_pos) if landlord_name_pos != -1 else float('inf')
                
                # Check if tenant name is closer to landlord section and vice versa
                if (tenant_to_landlord_section < tenant_to_tenant_section and 
                    landlord_to_tenant_section < landlord_to_landlord_section):
                    
                    st.warning("🔄 Detected name mix-up, swapping tenant and landlord")
                    
                    # Swap the information
                    temp_tenant = ContactInfo(
                        name=result.tenant.name,
                        phone=result.tenant.phone,
                        email=result.tenant.email
                    )
                    
                    result.tenant.name = result.landlord.name
                    result.tenant.phone = result.landlord.phone
                    result.tenant.email = result.landlord.email
                    
                    result.landlord.name = temp_tenant.name
                    result.landlord.phone = temp_tenant.phone
                    result.landlord.email = temp_tenant.email
                    
                    # Lower confidence slightly as we had to correct
                    result.confidence_score = max(0.1, result.confidence_score - 0.1)
                    result.extraction_method += " (Auto-corrected)"
                    
                    st.success(f"✅ Corrected - Tenant: {result.tenant.name}, Landlord: {result.landlord.name}")
            
        except Exception as e:
            st.warning(f"⚠️ Post-processing validation error: {str(e)}")
            # Don't crash on validation errors

def create_excel_report(results):
    """Create professional Excel report for download"""
    if not results:
        return None
    
    # Prepare data for DataFrame with only selected columns
    df_data = []
    for item in results:
        df_data.append({
            'File Name': item.file_name,
            'File Path': item.file_path,
            'Tenant Name': item.tenant.name,
            'Tenant Phone': item.tenant.phone,
            'Tenant Email': item.tenant.email,
            'Landlord Name': item.landlord.name,
            'Landlord Phone': item.landlord.phone,
            'Landlord Email': item.landlord.email,
            'Confidence Score': f"{item.confidence_score:.2f}"
        })
    
    df = pd.DataFrame(df_data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='F913 Extractions', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['F913 Extractions']
        
        # Format headers
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set column width with reasonable limits
            worksheet.column_dimensions[column_letter].width = min(max_length + 3, 60)
        
        # Add alternating row colors
        light_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = light_fill
    
    # Seek to the beginning of the stream
    output.seek(0)
    return output

# Set up the Streamlit app
st.set_page_config(
    page_title="F913 Document Extractor",
    page_icon="📄",
    layout="wide"
)

# Main app header
st.title("📄 F913 Document Extractor")
st.markdown("Extract tenant and landlord information from F913 Lease for Residential Property documents.")

# Sidebar for API key
with st.sidebar:
    st.header("🔑 API Key Setup")
    
    # Get API key from environment or user input
    api_key = os.getenv("GROQ_API_KEY")
    
    if not api_key:
        api_key = st.text_input("Enter your Groq API Key", type="password", 
                                help="Get your free API key from: https://console.groq.com/")
        
        if api_key:
            st.success("✅ API Key entered")
        else:
            st.warning("⚠️ Please enter a valid Groq API Key")
    else:
        st.success("✅ API Key loaded from environment")
        
    st.write("---")
    
    st.header("ℹ️ About")
    st.markdown("""
    This app extracts contact information from F913 lease documents.
    
    **Features:**
    - PDF text extraction
    - Advanced AI recognition
    - Regex fallback
    - Excel report generation
    
    **Contact info extracted:**
    - Tenant name, phone, email
    - Landlord name, phone, email
    """)
    
    st.write("---")
    
    # Select model 
    st.header("🤖 Model Selection")
    model_choice = st.radio(
        "Select Groq model:",
        ["premium", "fast", "instant"],
        index=0,
        help="Premium: Most accurate, Fast: Good balance, Instant: Quickest"
    )

# Initialize session state
if 'results' not in st.session_state:
    st.session_state.results = []

# Main content
st.header("Process F913 Documents")

# Option to choose between file upload or directory path
process_option = st.radio(
    "Choose processing method:",
    ["Upload Files", "Use Directory Path"],
    index=0
)

if process_option == "Upload Files":
    # Original file upload code
    uploaded_files = st.file_uploader("Choose PDF files", type="pdf", accept_multiple_files=True)
    
    # Process uploaded files
    if uploaded_files:
        # Initialize extractor
        extractor = GroqF913Extractor(api_key)
        
        if extractor.client:
            with st.spinner("🔍 Processing uploaded documents..."):
                batch_results = []
                
                for uploaded_file in uploaded_files:
                    # Create a container for each file
                    st.write("---")
                    file_container = st.container()
                    with file_container:
                        st.subheader(f"📄 Processing: {uploaded_file.name}")
                        
                        # Get file content
                        pdf_content = uploaded_file.read()
                        
                        # Process the file
                        result = extractor.extract_f913_data(pdf_content, uploaded_file.name)
                        
                        if result:
                            st.success(f"✅ Extraction completed | Confidence: {result.confidence_score:.2f}")
                            
                            # Show results in a nice format
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown("### 👤 Tenant Information")
                                st.markdown(f"**Name:** {result.tenant.name}")
                                st.markdown(f"**Phone:** {result.tenant.phone}")
                                st.markdown(f"**Email:** {result.tenant.email}")
                            
                            with col2:
                                st.markdown("### 🏠 Landlord Information")
                                st.markdown(f"**Name:** {result.landlord.name}")
                                st.markdown(f"**Phone:** {result.landlord.phone}")
                                st.markdown(f"**Email:** {result.landlord.email}")
                            
                            # Add to batch results
                            batch_results.append(result)
                        else:
                            st.error("❌ No data extracted from this file")
                
                # Store results in session state
                st.session_state.results = batch_results
        else:
            st.error("❌ Groq client not initialized. Please check your API key.")

else:  # Directory path option
    # Add directory path input
    directory_path = st.text_input("Enter root directory path:", 
                                  value=os.getcwd(),
                                  help="Enter the path to search for F913 PDFs (including all subdirectories)")
    
    # Add options for search behavior
    col1, col2 = st.columns(2)
    with col1:
        search_recursively = st.checkbox("Search subdirectories", value=True, 
                                      help="Search through all subfolders recursively")
    with col2:
        case_sensitive = st.checkbox("Case-sensitive search", value=False,
                                  help="Match F913/f913 case-sensitively")
    
    # Add a button to start processing
    if directory_path and st.button("Search and Process F913 Documents"):
        # Initialize extractor
        extractor = GroqF913Extractor(api_key)
        
        if extractor.client:
            # Check if directory exists
            if not os.path.exists(directory_path):
                st.error(f"❌ Directory not found: {directory_path}")
            else:
                # Find all PDF files containing "F913" in the directory (and subdirectories if selected)
                pdf_files = []
                pattern = "F913" if case_sensitive else "f913"
                
                with st.spinner("🔍 Searching for F913 PDF files..."):
                    if search_recursively:
                        # Walk through all subdirectories
                        for root, dirs, files in os.walk(directory_path):
                            for file in files:
                                if file.lower().endswith('.pdf'):
                                    if (case_sensitive and pattern in file) or (not case_sensitive and pattern.lower() in file.lower()):
                                        pdf_files.append(os.path.join(root, file))
                    else:
                        # Only search the specified directory (not subdirectories)
                        for file in os.listdir(directory_path):
                            if file.lower().endswith('.pdf'):
                                if (case_sensitive and pattern in file) or (not case_sensitive and pattern.lower() in file.lower()):
                                    pdf_files.append(os.path.join(directory_path, file))
                
                if not pdf_files:
                    st.warning(f"⚠️ No F913 PDF files found in {directory_path}")
                else:
                    st.info(f"🔍 Found {len(pdf_files)} F913 PDF files")
                    
                    # Show file list in an expander
                    with st.expander("📋 Found Files"):
                        for i, file_path in enumerate(pdf_files, 1):
                            st.text(f"{i}. {file_path}")
                    
                    # Process all PDF files
                    with st.spinner(f"Processing {len(pdf_files)} PDF files..."):
                        batch_results = []
                        progress_bar = st.progress(0)
                        
                        for i, pdf_path in enumerate(pdf_files):
                            # Create a container for each file
                            file_container = st.container()
                            with file_container:
                                file_name = os.path.basename(pdf_path)
                                st.subheader(f"📄 Processing: {file_name}")
                                st.caption(f"Path: {pdf_path}")
                                
                                try:
                                    # Read PDF file
                                    with open(pdf_path, 'rb') as file:
                                        pdf_content = file.read()
                                    
                                    # Process the file
                                    result = extractor.extract_f913_data(pdf_content, file_name)
                                    
                                    if result:
                                        # Store the full path
                                        result.file_path = pdf_path
                                        
                                        st.success(f"✅ Extraction completed | Confidence: {result.confidence_score:.2f}")
                                        
                                        # Show results in a nice format
                                        col1, col2 = st.columns(2)
                                        
                                        with col1:
                                            st.markdown("### 👤 Tenant Information")
                                            st.markdown(f"**Name:** {result.tenant.name}")
                                            st.markdown(f"**Phone:** {result.tenant.phone}")
                                            st.markdown(f"**Email:** {result.tenant.email}")
                                        
                                        with col2:
                                            st.markdown("### 🏠 Landlord Information")
                                            st.markdown(f"**Name:** {result.landlord.name}")
                                            st.markdown(f"**Phone:** {result.landlord.phone}")
                                            st.markdown(f"**Email:** {result.landlord.email}")
                                        
                                        # Add to batch results
                                        batch_results.append(result)
                                    else:
                                        st.error("❌ No data extracted from this file")
                                        
                                except Exception as e:
                                    st.error(f"❌ Error processing {file_name}: {str(e)}")
                            
                            # Update progress
                            progress_bar.progress((i + 1) / len(pdf_files))
                        
                        # Store results in session state
                        st.session_state.results = batch_results
                        st.success(f"✅ Processed {len(batch_results)} files successfully")
        else:
            st.error("❌ Groq client not initialized. Please check your API key.")

# Show results summary and download option
if st.session_state.results:
    st.write("---")
    st.header("📊 Extraction Results Summary")
    
    # Create a dataframe for display
    display_data = []
    for result in st.session_state.results:
        display_data.append({
            "File Name": result.file_name,
            "Tenant Name": result.tenant.name,
            "Landlord Name": result.landlord.name,
            "Confidence": f"{result.confidence_score:.2f}"
        })
    
    df_display = pd.DataFrame(display_data)
    st.dataframe(df_display)
    
    # Create Excel report for download
    excel_data = create_excel_report(st.session_state.results)
    if excel_data:
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name=f"F913_Extractions_{now}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Demo section with instructions
else:
    st.write("---")
    if api_key:
        st.info("👆 Upload F913 PDF documents or specify a directory to extract tenant and landlord information")
    else:
        st.warning("⚠️ Please enter your Groq API Key in the sidebar to start processing")

    # Sample instructions
    with st.expander("📘 How to use this app"):
        st.markdown("""
        1. **Enter your Groq API Key** in the sidebar (if not already loaded)
        2. **Choose processing method**:
           - Upload PDF files directly
           - Or specify a directory path where your PDF files are stored
        3. **Wait for processing** - this may take a few seconds per file
        4. **Review the extracted information** for each document
        5. **Download the Excel report** with all extracted data
        
        The app uses Groq's AI models to intelligently extract contact information from lease documents.
        If the AI extraction fails, a fallback regex-based extraction is attempted.
        """)