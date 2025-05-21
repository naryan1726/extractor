#!/usr/bin/env python3
"""
Groq-Optimized F913 Document Extraction Tool
Uses latest Groq models (Jan 2025) for maximum speed and accuracy
"""

import os
import re
import json
import pandas as pd
import traceback
import time
from datetime import datetime
from typing import Dict, List, Optional
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()  # This loads the .env file
    print("✅ .env file loaded successfully")
except ImportError:
    print("⚠️ python-dotenv not installed. Install with: pip install python-dotenv")
except Exception as e:
    print(f"⚠️ Could not load .env file: {e}")

# For PDF processing
try:
    import pdfplumber
    PDF_LIBRARY = "pdfplumber"
except ImportError:
    try:
        from PyPDF2 import PdfReader
        PDF_LIBRARY = "PyPDF2"
    except ImportError:
        PDF_LIBRARY = None

# Groq API
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

@dataclass
class ContactInfo:
    """Contact information data structure"""
    name: str = ""
    phone: str = ""
    email: str = ""

@dataclass
class F913ExtractionResult:
    """F913 extraction result with metadata"""
    file_name: str
    file_path: str
    tenant: ContactInfo
    landlord: ContactInfo
    extraction_method: str = "Unknown"
    model_used: str = ""
    confidence_score: float = 0.0
    processing_time: float = 0.0
    processed_datetime: str = ""  # Added timestamp field

class GroqF913Extractor:
    """High-performance F913 extractor using latest Groq models"""
    
    # Latest Groq models (Jan 2025)
    MODELS = {
        "premium": "llama-3.3-70b-versatile",      # Best accuracy, 128K context
        "fast": "llama-3.3-70b-specdec",          # Fastest 70B, speculatuve decoding
        "instant": "llama-3.1-8b-instant",        # Quick responses, lower cost
        "fallback": "llama-3.1-70b-versatile"     # Backup option
    }
    
    def __init__(self, api_key: Optional[str] = None):
        """Initialize with Groq API key"""
        # Check environment variable first
        self.api_key = api_key or os.environ.get("GROQ_API_KEY") or os.getenv("GROQ_API_KEY")
        self.client = None
        self.default_model = self.MODELS["premium"]
        
        if not self.api_key:
            print("❌ GROQ_API_KEY not found in environment variables!")
            print("💡 Get your free API key from: https://console.groq.com/")
            print("💡 Set environment variable:")
            print("   Windows: set GROQ_API_KEY=your_key_here")
            print("   macOS/Linux: export GROQ_API_KEY=your_key_here")
            print("   Or add to .env file: GROQ_API_KEY=your_key_here")
            return
        
        if not GROQ_AVAILABLE:
            print("❌ Groq library not installed!")
            print("💡 Install with: pip install groq")
            return
        
        try:
            self.client = Groq(api_key=self.api_key)
            self._test_connection()
        except Exception as e:
            print(f"❌ Groq initialization failed: {str(e)}")
            self.client = None
    
    def _test_connection(self):
        """Test Groq API connection"""
        try:
            # Quick test with smallest model
            response = self.client.chat.completions.create(
                model=self.MODELS["instant"],
                messages=[{"role": "user", "content": "Hi"}],
                max_tokens=5
            )
            print("✅ Groq API connected successfully")
            print(f"🚀 Using models: {', '.join(self.MODELS.values())}")
        except Exception as e:
            print(f"⚠️ Groq API test failed: {str(e)}")
            raise
    
    def search_for_f913_documents(self, root_dir: str) -> List[F913ExtractionResult]:
        """Search and process F913 documents in directory"""
        if not self.client:
            print("❌ Groq client not available")
            return []
        
        if not PDF_LIBRARY:
            print("❌ PDF processing library not found")
            print("💡 Install with: pip install pdfplumber")
            return []
        
        print(f"\n🔍 Searching for F913 documents in: {root_dir}")
        print("="*70)
        
        results = []
        total_files = 0
        start_time = time.time()
        
        # Walk through directory structure
        for root, dirs, files in os.walk(root_dir):
            f913_files = [f for f in files if 'F913' in f.upper() and f.lower().endswith('.pdf')]
            
            if f913_files:
                print(f"\n📁 Found {len(f913_files)} F913 files in: {os.path.basename(root)}")
                
                for pdf_file in f913_files:
                    total_files += 1
                    file_path = os.path.join(root, pdf_file)
                    
                    print(f"\n📄 [{total_files}] Processing: {pdf_file}")
                    
                    result = self.extract_f913_data(file_path)
                    if result:
                        # Display results
                        success = result.tenant.name or result.landlord.name
                        status = "✅" if success else "⚠️"
                        
                        print(f"  {status} Model: {result.model_used} | "
                              f"Time: {result.processing_time:.1f}s | "
                              f"Confidence: {result.confidence_score:.1f}")
                        
                        if success:
                            print(f"    👤 Tenant: {result.tenant.name} | {result.tenant.phone} | {result.tenant.email}")
                            print(f"    🏠 Landlord: {result.landlord.name} | {result.landlord.phone} | {result.landlord.email}")
                        
                        results.append(result)
                    else:
                        print("  ❌ Extraction failed")
        
        # Summary
        total_time = time.time() - start_time
        successful = len([r for r in results if r.tenant.name or r.landlord.name])
        
        print("\n" + "="*70)
        print(f"📊 PROCESSING COMPLETE")
        print("="*70)
        print(f"Total files processed: {total_files}")
        print(f"Successful extractions: {successful}")
        print(f"Total processing time: {total_time:.1f} seconds")
        print(f"Average per file: {total_time/max(total_files, 1):.1f} seconds")
        print("="*70)
        
        return results
    
    def extract_f913_data(self, pdf_path: str) -> Optional[F913ExtractionResult]:
        """Extract data from single F913 PDF using Groq"""
        start_time = time.time()
        
        # Create result with timestamp
        processed_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        result = F913ExtractionResult(
            file_name=os.path.basename(pdf_path),
            file_path=pdf_path,
            tenant=ContactInfo(),
            landlord=ContactInfo(),
            processed_datetime=processed_datetime
        )
        
        try:
            # Extract PDF text
            signature_text, full_text = self._extract_pdf_text(pdf_path)
            
            if not signature_text and not full_text:
                print("    ❌ No text extracted from PDF")
                return None
            
            # Use signature page if available, otherwise full text
            text_to_analyze = signature_text or full_text
            
            # Truncate text for API efficiency
            if len(text_to_analyze) > 8000:
                text_to_analyze = text_to_analyze[:8000] + "..."
            
            # Try extraction with different models
            for model_type, model_id in self.MODELS.items():
                try:
                    print(f"    🤖 Trying {model_type} model: {model_id}")
                    
                    extracted_data = self._extract_with_groq_model(
                        text_to_analyze, model_id, model_type
                    )
                    
                    if extracted_data:
                        # Update result
                        result.tenant = ContactInfo(**extracted_data.get("tenant", {}))
                        result.landlord = ContactInfo(**extracted_data.get("landlord", {}))
                        result.extraction_method = "Groq API"
                        result.model_used = f"{model_type} ({model_id})"
                        result.confidence_score = extracted_data.get("confidence", 0.8)
                        result.processing_time = time.time() - start_time
                        
                        # Validate and correct potential mix-ups
                        self._post_process_validation(result, text_to_analyze)
                        
                        # Check if we got useful data
                        if result.tenant.name or result.landlord.name:
                            return result
                
                except Exception as e:
                    print(f"    ⚠️ {model_type} model failed: {str(e)}")
                    continue
            
            # If all models failed, try regex fallback
            print("    🔍 Falling back to regex extraction")
            regex_result = self._extract_with_regex(text_to_analyze, result)
            if regex_result:
                regex_result.processing_time = time.time() - start_time
                return regex_result
            
            return None
            
        except Exception as e:
            print(f"    ❌ Error processing {pdf_path}: {str(e)}")
            return None
    
    def _extract_pdf_text(self, pdf_path: str) -> tuple:
        """Extract text from PDF, prioritizing signature page"""
        signature_text = ""
        full_text = ""
        
        try:
            if PDF_LIBRARY == "pdfplumber":
                with pdfplumber.open(pdf_path) as pdf:
                    pages = pdf.pages
                    
                    # Check last page for signature
                    if pages:
                        last_page_text = pages[-1].extract_text() or ""
                        if any(marker in last_page_text for marker in 
                               ["IN WITNESS WHEREOF", "Tenant's Signature", "Landlord's Signature"]):
                            signature_text = last_page_text
                            print("      📋 Found signature page")
                    
                    # Get full text as fallback
                    for page in pages:
                        page_text = page.extract_text() or ""
                        full_text += page_text + "\n"
                        
                        # Look for signature page in all pages if not found
                        if not signature_text and "IN WITNESS WHEREOF" in page_text:
                            signature_text = page_text
            
            elif PDF_LIBRARY == "PyPDF2":
                with open(pdf_path, 'rb') as file:
                    reader = PdfReader(file)
                    for page in reader.pages:
                        page_text = page.extract_text() or ""
                        full_text += page_text + "\n"
                        
                        if "IN WITNESS WHEREOF" in page_text and not signature_text:
                            signature_text = page_text
            
            return signature_text.strip(), full_text.strip()
            
        except Exception as e:
            print(f"      ⚠️ PDF extraction error: {str(e)}")
            return "", ""
    
    def _extract_with_groq_model(self, text: str, model_id: str, model_type: str) -> Optional[dict]:
        """Extract data using specific Groq model"""
        
        # Adjust prompt based on model capability
        if model_type in ["premium", "fast"]:
            # Detailed prompt for powerful models with improved context separation
            prompt = f"""You are an expert at extracting contact information from F913 Lease for Residential Property documents.

TASK: Analyze this F913 document text and extract tenant and landlord contact information.

IMPORTANT INSTRUCTIONS:
1. Look for the signature page section with "IN WITNESS WHEREOF"
2. Find TWO distinct sections:
   - "Tenant's Signature" section (left side) - contains TENANT information
   - "Landlord's Signature" section (right side) - contains LANDLORD information
3. Each section has "Print or Type Name", phone number, and email address fields
4. DO NOT mix up the two parties - tenant info goes to tenant, landlord info goes to landlord
5. If you see names like "Sam Kumar" and "Apuroop", determine which is tenant vs landlord based on their position in the document

CRITICAL: Pay attention to the document structure - there are separate fields for each party.

Return ONLY a valid JSON object in this exact format:
{{
    "tenant": {{
        "name": "Full Name Here",
        "phone": "1234567890",
        "email": "email@domain.com"
    }},
    "landlord": {{
        "name": "Full Name Here", 
        "phone": "1234567890",
        "email": "email@domain.com"
    }},
    "confidence": 0.95
}}

Rules:
- Use empty strings "" for any missing information
- Phone numbers should be 10 digits only
- DO NOT switch tenant and landlord information
- First name/phone/email usually belongs to tenant, second to landlord
- Confidence should be 0.0 to 1.0
- Return ONLY the JSON, no explanation

Document text:
{text[:4000]}"""
        
        else:
            # Simplified prompt for smaller models
            prompt = f"""Extract tenant and landlord info from F913 lease document.

IMPORTANT: Don't mix up tenant and landlord info. First person is usually tenant, second is landlord.

Return JSON only:
{{
    "tenant": {{"name": "", "phone": "", "email": ""}},
    "landlord": {{"name": "", "phone": "", "email": ""}},
    "confidence": 0.8
}}

Document: {text[:2500]}"""
        
        try:
            # Configure based on model type
            if model_type == "instant":
                max_tokens = 300
                temperature = 0.1
            elif model_type == "fast":
                max_tokens = 400  
                temperature = 0.05
            else:  # premium
                max_tokens = 500
                temperature = 0.02
            
            response = self.client.chat.completions.create(
                model=model_id,
                messages=[
                    {
                        "role": "system", 
                        "content": "You are a precise document data extraction specialist. Return only valid JSON responses."
                    },
                    {
                        "role": "user", 
                        "content": prompt
                    }
                ],
                max_tokens=max_tokens,
                temperature=temperature,
                top_p=0.9
            )
            
            content = response.choices[0].message.content.strip()
            
            # Clean JSON response
            content = self._clean_json_response(content)
            
            # Parse JSON
            data = json.loads(content)
            
            # Validate structure
            if self._validate_extraction_data(data):
                return data
            else:
                print(f"      ⚠️ Invalid data structure from {model_type}")
                return None
                
        except json.JSONDecodeError as e:
            print(f"      ⚠️ JSON parse error from {model_type}: {str(e)}")
            # Try to extract JSON from response
            return self._extract_json_from_text(content)
        except Exception as e:
            print(f"      ⚠️ {model_type} API error: {str(e)}")
            return None
    
    def _clean_json_response(self, content: str) -> str:
        """Clean and extract JSON from model response"""
        # Remove markdown formatting
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        
        # Find JSON object
        start = content.find('{')
        end = content.rfind('}') + 1
        
        if start >= 0 and end > start:
            return content[start:end]
        
        return content.strip()
    
    def _extract_json_from_text(self, text: str) -> Optional[dict]:
        """Extract JSON from text using regex"""
        try:
            # Find JSON-like structure
            json_pattern = r'\{[^{}]*\{[^{}]*\}[^{}]*\{[^{}]*\}[^{}]*\}'
            match = re.search(json_pattern, text, re.DOTALL)
            
            if match:
                return json.loads(match.group())
            
            # Fallback: try to parse any {.*} structure
            simple_pattern = r'\{.*\}'
            match = re.search(simple_pattern, text, re.DOTALL)
            
            if match:
                return json.loads(match.group())
            
            return None
        except:
            return None
    
    def _validate_extraction_data(self, data: dict) -> bool:
        """Validate extracted data structure"""
        try:
            return (
                isinstance(data, dict) and
                "tenant" in data and
                "landlord" in data and
                isinstance(data["tenant"], dict) and
                isinstance(data["landlord"], dict)
            )
        except:
            return False
    
    def _extract_with_regex(self, text: str, result: F913ExtractionResult) -> Optional[F913ExtractionResult]:
        """Fallback regex extraction with improved party separation"""
        try:
            # Clean and normalize text
            normalized_text = re.sub(r'\s+', ' ', text)
            
            # Try to split the signature section into tenant and landlord parts
            tenant_section = ""
            landlord_section = ""
            
            # Look for section separators
            if "Tenant's Signature" in text and "Landlord's Signature" in text:
                # Split by landlord signature to separate sections
                parts = text.split("Landlord's Signature")
                if len(parts) >= 2:
                    tenant_section = parts[0]  # Everything before landlord section
                    landlord_section = "Landlord's Signature" + parts[1]  # Landlord section
                    print("      📋 Found separate tenant and landlord sections")
            
            # F913-specific name patterns with section awareness
            if tenant_section and landlord_section:
                # Extract from separate sections
                tenant_name_patterns = [
                    r'Tenant[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                    r'1\s+Tenant[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print',
                ]
                
                landlord_name_patterns = [
                    r'Landlord[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                    r'1\s+Landlord[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print',
                ]
                
                # Extract tenant name from tenant section
                for pattern in tenant_name_patterns:
                    match = re.search(pattern, tenant_section, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.tenant.name = name
                            print(f"      👤 Tenant name from section: {name}")
                            break
                
                # Extract landlord name from landlord section
                for pattern in landlord_name_patterns:
                    match = re.search(pattern, landlord_section, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.landlord.name = name
                            print(f"      🏠 Landlord name from section: {name}")
                            break
                
                # Extract phones with context
                tenant_phones = re.findall(r'(\d{10})', tenant_section)
                landlord_phones = re.findall(r'(\d{10})', landlord_section)
                
                if tenant_phones:
                    result.tenant.phone = tenant_phones[0]
                    print(f"      📞 Tenant phone from section: {tenant_phones[0]}")
                
                if landlord_phones:
                    result.landlord.phone = landlord_phones[0]
                    print(f"      📞 Landlord phone from section: {landlord_phones[0]}")
                
                # Extract emails with context
                tenant_emails = re.findall(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', tenant_section)
                landlord_emails = re.findall(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', landlord_section)
                
                if tenant_emails:
                    result.tenant.email = tenant_emails[0]
                    print(f"      📧 Tenant email from section: {tenant_emails[0]}")
                
                if landlord_emails:
                    result.landlord.email = landlord_emails[0]
                    print(f"      📧 Landlord email from section: {landlord_emails[0]}")
            
            else:
                # Fallback to original patterns but with better logic
                print("      ⚠️ Could not separate sections, using positional extraction")
                
                patterns = {
                    'tenant_name': [
                        r'1\s+Tenant[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                        r'Tenant[\'s]*\s+Signature\s+([A-Za-z\s]{3,40}?)\s+Print\s+or\s+Type\s+Name',
                    ],
                    'landlord_name': [
                        r'1\s+Landlord[\'s]*\s+Signature\s+([A-Za-z\s]+?)\s+Print\s+or\s+Type\s+Name',
                        r'Landlord[\'s]*\s+Signature\s+([A-Za-z\s]{3,40}?)\s+Print\s+or\s+Type\s+Name',
                    ]
                }
                
                # Extract names using order-aware patterns
                for pattern_list in patterns['tenant_name']:
                    match = re.search(pattern_list, normalized_text, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.tenant.name = name
                            print(f"      👤 Tenant name: {name}")
                            break
                
                for pattern_list in patterns['landlord_name']:
                    match = re.search(pattern_list, normalized_text, re.IGNORECASE | re.DOTALL)
                    if match:
                        name = match.group(1).strip()
                        if 2 < len(name) < 50:
                            result.landlord.name = name
                            print(f"      🏠 Landlord name: {name}")
                            break
                
                # Extract phones and emails with positional logic
                all_phones = re.findall(r'(\d{10})', normalized_text)
                all_emails = re.findall(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', normalized_text)
                
                # Assign based on order (first phone/email to tenant, second to landlord)
                if all_phones:
                    if len(all_phones) >= 1:
                        result.tenant.phone = all_phones[0]
                        print(f"      📞 Tenant phone (position 1): {all_phones[0]}")
                    if len(all_phones) >= 2:
                        result.landlord.phone = all_phones[1]
                        print(f"      📞 Landlord phone (position 2): {all_phones[1]}")
                
                if all_emails:
                    if len(all_emails) >= 1:
                        result.tenant.email = all_emails[0]
                        print(f"      📧 Tenant email (position 1): {all_emails[0]}")
                    if len(all_emails) >= 2:
                        result.landlord.email = all_emails[1]
                        print(f"      📧 Landlord email (position 2): {all_emails[1]}")
            
            # Validation: Check if we mixed up names (common names heuristic)
            self._validate_and_correct_assignment(result)
            
            result.extraction_method = "Enhanced Regex"
            result.model_used = "Pattern Matching with Section Separation"
            result.confidence_score = 0.7 if (result.tenant.name or result.landlord.name) else 0.3
            
            return result if (result.tenant.name or result.landlord.name or result.tenant.phone or result.tenant.email) else None
            
        except Exception as e:
            print(f"      ⚠️ Regex extraction error: {str(e)}")
            return None
    
    def _validate_and_correct_assignment(self, result: F913ExtractionResult):
        """Validate and correct tenant/landlord assignment if needed"""
        try:
            # Simple heuristics to detect if assignments might be wrong
            # This is based on common patterns in F913 documents
            
            tenant_name = result.tenant.name.lower() if result.tenant.name else ""
            landlord_name = result.landlord.name.lower() if result.landlord.name else ""
            
            # Check if we have both names
            if tenant_name and landlord_name:
                # If landlord name appears to be a first name only while tenant has full name,
                # they might be swapped
                tenant_parts = tenant_name.split()
                landlord_parts = landlord_name.split()
                
                # Heuristic: If "landlord" has only one word and "tenant" has multiple,
                # and landlord word appears in tenant's name, they might be swapped
                if (len(landlord_parts) == 1 and len(tenant_parts) > 1 and 
                    landlord_parts[0] in tenant_parts):
                    print("      ⚠️ Detected possible name mix-up, checking...")
                    
                    # Don't auto-correct as it might be wrong, just flag
                    result.confidence_score = max(0.1, result.confidence_score - 0.2)
                    print("      📝 Lowered confidence due to potential mix-up")
            
        except Exception:
            # If validation fails, don't crash
            pass
    
    def _post_process_validation(self, result: F913ExtractionResult, original_text: str):
        """Post-process validation to detect and fix common mix-ups"""
        try:
            # Check for obvious mix-ups by analyzing the original text structure
            if not (result.tenant.name and result.landlord.name):
                return  # Can't validate if we don't have both names
            
            # Look for positional clues in the original text
            tenant_name_pos = original_text.find(result.tenant.name)
            landlord_name_pos = original_text.find(result.landlord.name)
            
            # Look for section indicators
            tenant_section_pos = original_text.find("Tenant's Signature")
            landlord_section_pos = original_text.find("Landlord's Signature")
            
            if tenant_section_pos != -1 and landlord_section_pos != -1:
                # Check if names are in wrong sections
                tenant_to_tenant_section = abs(tenant_name_pos - tenant_section_pos) if tenant_name_pos != -1 else float('inf')
                tenant_to_landlord_section = abs(tenant_name_pos - landlord_section_pos) if tenant_name_pos != -1 else float('inf')
                
                landlord_to_tenant_section = abs(landlord_name_pos - tenant_section_pos) if landlord_name_pos != -1 else float('inf')
                landlord_to_landlord_section = abs(landlord_name_pos - landlord_section_pos) if landlord_name_pos != -1 else float('inf')
                
                # Check if tenant name is closer to landlord section and vice versa
                if (tenant_to_landlord_section < tenant_to_tenant_section and 
                    landlord_to_tenant_section < landlord_to_landlord_section):
                    
                    print("      🔄 Detected name mix-up, swapping tenant and landlord")
                    
                    # Swap the information
                    temp_tenant = ContactInfo(
                        name=result.tenant.name,
                        phone=result.tenant.phone,
                        email=result.tenant.email
                    )
                    
                    result.tenant.name = result.landlord.name
                    result.tenant.phone = result.landlord.phone
                    result.tenant.email = result.landlord.email
                    
                    result.landlord.name = temp_tenant.name
                    result.landlord.phone = temp_tenant.phone
                    result.landlord.email = temp_tenant.email
                    
                    # Lower confidence slightly as we had to correct
                    result.confidence_score = max(0.1, result.confidence_score - 0.1)
                    result.extraction_method += " (Auto-corrected)"
                    
                    print(f"      ✅ Corrected - Tenant: {result.tenant.name}, Landlord: {result.landlord.name}")
            
        except Exception as e:
            print(f"      ⚠️ Post-processing validation error: {str(e)}")
            # Don't crash on validation errors
    
    def create_excel_report(self, data: List[F913ExtractionResult], output_file: str) -> bool:
        """Create professional Excel report with timestamp"""
        if not data:
            print("❌ No data to export")
            return False
        
        try:
            print(f"\n📊 Creating Excel report with {len(data)} records...")
            
            # Prepare data for DataFrame
            df_data = []
            for item in data:
                df_data.append({
                    'File Name': item.file_name,
                    'File Path': item.file_path,
                    'Tenant Name': item.tenant.name,
                    'Tenant Phone': item.tenant.phone,
                    'Tenant Email': item.tenant.email,
                    'Landlord Name': item.landlord.name,
                    'Landlord Phone': item.landlord.phone,
                    'Landlord Email': item.landlord.email,
                    'Confidence Score': f"{item.confidence_score:.2f}"
                })
            
            df = pd.DataFrame(df_data)
            
            # Create Excel with formatting
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='F913 Extractions', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['F913 Extractions']
                
                # Header formatting
                header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    # Set column width with reasonable limits
                    worksheet.column_dimensions[column_letter].width = min(max_length + 3, 60)
                
                # Add alternating row colors
                light_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                for row_num in range(2, len(df) + 2):
                    if row_num % 2 == 0:
                        for col_num in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = light_fill
            
            print(f"✅ Excel report created: {output_file}")
            
            # Summary statistics
            successful = len([r for r in data if r.tenant.name or r.landlord.name])
            avg_time = sum(r.processing_time for r in data) / len(data)
            avg_confidence = sum(r.confidence_score for r in data) / len(data)
            
            print(f"📈 Report Summary:")
            print(f"   • Total documents: {len(data)}")
            print(f"   • Successful extractions: {successful}")
            print(f"   • Average processing time: {avg_time:.1f}s")
            print(f"   • Average confidence: {avg_confidence:.2f}")
            
            return True
            
        except Exception as e:
            print(f"❌ Excel creation failed: {str(e)}")
            return False

def print_setup_instructions():
    """Print setup instructions"""
    print("\n" + "="*70)
    print("🛠️  GROQ SETUP INSTRUCTIONS")
    print("="*70)
    print("1. Get FREE Groq API Key:")
    print("   • Visit: https://console.groq.com/")
    print("   • Sign up (completely free)")
    print("   • Create API key")
    print("   • Copy the key (starts with 'gsk_')")
    
    print("\n2. Set Environment Variable:")
    print("   • Windows: set GROQ_API_KEY=your_key_here")
    print("   • macOS/Linux: export GROQ_API_KEY=your_key_here")
    print("   • Or add to .env file: GROQ_API_KEY=your_key_here")
    
    print("\n3. Install Dependencies:")
    print("   • pip install groq pdfplumber pandas openpyxl")
    
    print("\n4. Run the Tool:")
    print("   • python groq_optimized_f913_extractor.py")
    
    print("\n💡 Free Tier Limits:")
    print("   • 14,400 requests per day")
    print("   • More than enough for document processing!")
    
    print("\n🔍 Checking Environment Variables:")
    groq_key = os.environ.get("GROQ_API_KEY") or os.getenv("GROQ_API_KEY")
    if groq_key:
        print(f"   ✅ GROQ_API_KEY found: {groq_key[:10]}...")
    else:
        print("   ❌ GROQ_API_KEY not found in environment")
    print("="*70)

def get_timestamped_filename(base_name: str) -> str:
    """Generate filename with timestamp"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name_part, ext = os.path.splitext(base_name)
    return f"{name_part}_{timestamp}{ext}"

def main():
    """Main application entry point"""
    try:
        print("🚀 GROQ-OPTIMIZED F913 DOCUMENT EXTRACTOR 🚀")
        print("Using Latest Groq Models (January 2025)")
        print("="*70)
        
        # Check for API key first with better detection
        api_key = os.environ.get("GROQ_API_KEY") or os.getenv("GROQ_API_KEY")
        
        if not api_key:
            print("\n❌ GROQ_API_KEY not found in environment variables!")
            print_setup_instructions()
            
            # Ask user for API key as fallback
            api_key = input("\n🔑 Enter your Groq API key manually (or press Enter to exit): ").strip()
            
            if not api_key:
                print("\n👋 Exiting. Please set GROQ_API_KEY environment variable and try again.")
                return
            
            print("⚠️ Note: For permanent solution, set GROQ_API_KEY as environment variable")
        else:
            print(f"✅ Found GROQ_API_KEY in environment: {api_key[:10]}...")
        
        # Initialize extractor
        extractor = GroqF913Extractor(api_key)
        if not extractor.client:
            print("❌ Failed to initialize Groq client")
            print_setup_instructions()
            return
        
        # Get processing choice
        current_dir = os.getcwd()
        print(f"\n📂 Current directory: {current_dir}")
        
        choice = input("\nChoose option:\n"
                      "(1) Search directory for F913 files\n"
                      "(2) Process single F913 file\n"
                      "(3) Show setup instructions\n"
                      "\nChoice (1/2/3): ").strip()
        
        if choice == "3":
            print_setup_instructions()
            return
        
        elif choice == "1":
            # Directory processing
            root_dir = input(f"\n📁 Enter directory path [default: {current_dir}]: ").strip()
            if not root_dir:
                root_dir = current_dir
            
            root_dir = root_dir.strip('"').replace('\\\\', '\\')
            
            if not os.path.exists(root_dir):
                print(f"❌ Directory not found: {root_dir}")
                return
            
            results = extractor.search_for_f913_documents(root_dir)
            
        elif choice == "2":
            # Single file processing
            file_path = input("\n📄 Enter F913 PDF file path: ").strip()
            file_path = file_path.strip('"').replace('\\\\', '\\')
            
            if not os.path.exists(file_path):
                print(f"❌ File not found: {file_path}")
                return
            
            print("\n🔄 Processing single file...")
            result = extractor.extract_f913_data(file_path)
            results = [result] if result else []
            
            if result:
                print(f"✅ Extraction completed:")
                print(f"   Model: {result.model_used}")
                print(f"   Time: {result.processing_time:.1f}s")
                print(f"   Confidence: {result.confidence_score:.2f}")
                print(f"   Processed: {result.processed_datetime}")
            else:
                print("❌ No data extracted")
        
        else:
            print("❌ Invalid choice")
            return
        
        # Process results
        if results:
            successful = [r for r in results if r.tenant.name or r.landlord.name]
            
            if successful:
                print(f"\n📋 EXTRACTION RESULTS")
                print("-" * 50)
                
                for i, doc in enumerate(successful, 1):
                    print(f"\n[{i}] {doc.file_name}")
                    print(f"    👤 Tenant: {doc.tenant.name} | {doc.tenant.phone} | {doc.tenant.email}")
                    print(f"    🏠 Landlord: {doc.landlord.name} | {doc.landlord.phone} | {doc.landlord.email}")
                    print(f"    ⏰ Processed: {doc.processed_datetime}")
                
                # Create Excel report with timestamp
                current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
                default_output = os.path.join(os.getcwd(), f"F913_Groq_Extractions_{current_time}.xlsx")
                
                output_choice = input(f"\n💾 Save Excel report?\n"
                                    f"(1) Use timestamped filename: F913_Groq_Extractions_{current_time}.xlsx\n"
                                    f"(2) Enter custom filename\n"
                                    f"(3) Skip report\n"
                                    f"Choice (1/2/3): ").strip()
                
                if output_choice == "1":
                    output_file = default_output
                elif output_choice == "2":
                    custom_name = input("Enter filename (with .xlsx extension): ").strip()
                    if not custom_name.endswith('.xlsx'):
                        custom_name += '.xlsx'
                    
                    # Add timestamp to custom name
                    name_part, ext = os.path.splitext(custom_name)
                    output_file = os.path.join(os.getcwd(), f"{name_part}_{current_time}{ext}")
                elif output_choice == "3":
                    print("📊 Report creation skipped")
                    output_file = None
                else:
                    print("⚠️ Invalid choice, using default timestamped filename")
                    output_file = default_output
                
                if output_file:
                    output_file = output_file.strip('"').replace('\\\\', '\\')
                    success = extractor.create_excel_report(results, output_file)
                    if success:
                        print(f"🎯 Excel report saved with timestamp: {os.path.basename(output_file)}")
            
            else:
                print("\n❌ No successful extractions")
        else:
            print("\n❌ No documents processed")
        
        print("\n🎉 Processing complete!")
        print(f"⏰ Session completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        input("\n⏸️ Press Enter to exit...")
        
    except KeyboardInterrupt:
        print("\n\n⏹️ Process interrupted by user")
    except Exception as e:
        print(f"\n❌ Unexpected error: {str(e)}")
        print(traceback.format_exc())
        input("\n⏸️ Press Enter to exit...")

if __name__ == "__main__":
    main()